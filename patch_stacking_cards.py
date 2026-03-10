"""
Patches the stacking-counter cards and Nash Thornton into the Excel spreadsheet.
Run from the project root:
    python patch_stacking_cards.py

Cards updated:
  00011 Desmond Wallace  (OL)  - StackStat: +2 run per consecutive run, uncapped
  00026 Desean Collins   (OL)  - StackStat: +1 run per consecutive run, max +5
  00059 Damien Vaughn    (RB)  - StackStat: +1 run per consecutive run, max +3
  00048 Nash Thornton    (QB)  - DiscardCard: opponent discards on incomplete pass
  00099 Ibrahim Lewis    (DB)  - AddStat: +3 short/deep coverage on 2nd consecutive pass
  00129 Xavier Morris    (LB)  - AddStat: +3 run coverage on 2nd consecutive run
"""

import openpyxl
import os

XLSX_PATH = os.path.join("Assets", "Resources", "First and Long - card_exporter.xlsx")

# Column layout (1-based, matching create_card_spreadsheet.py)
# Cols 15-23 = ability 1;  Cols 24-32 = ability 2
COL_A1_TRIGGER    = 15
COL_A1_TARGET     = 16
COL_A1_FAIL_EVENT = 17
COL_A1_COND       = 18
COL_A1_COND_VAL   = 19
COL_A1_COND2      = 20
COL_A1_COND2_VAL  = 21
COL_A1_EFFECT     = 22
COL_A1_EFFECT_VAL = 23

COL_A2_TRIGGER    = 24
COL_A2_TARGET     = 25
COL_A2_FAIL_EVENT = 26
COL_A2_COND       = 27
COL_A2_COND_VAL   = 28
COL_A2_COND2      = 29
COL_A2_COND2_VAL  = 30
COL_A2_EFFECT     = 31
COL_A2_EFFECT_VAL = 32

# Patches: card_id -> {col_index: value, ...}
# Use None to clear a cell.
PATCHES = {
    # ── Desmond Wallace (OL): +2 run per consecutive run, uncapped ──────────
    "00011": {
        COL_A1_TRIGGER:    "OnRun",
        COL_A1_TARGET:     "Self",
        COL_A1_FAIL_EVENT: None,
        COL_A1_COND:       None,
        COL_A1_COND_VAL:   None,
        COL_A1_COND2:      None,
        COL_A1_COND2_VAL:  None,
        COL_A1_EFFECT:     "StackStat",
        COL_A1_EFFECT_VAL: "RunBonus|Run|2|0",
        # clear ability 2
        COL_A2_TRIGGER:    None,
        COL_A2_TARGET:     None,
        COL_A2_FAIL_EVENT: None,
        COL_A2_COND:       None,
        COL_A2_COND_VAL:   None,
        COL_A2_COND2:      None,
        COL_A2_COND2_VAL:  None,
        COL_A2_EFFECT:     None,
        COL_A2_EFFECT_VAL: None,
    },

    # ── Desean Collins (OL): +1 run per consecutive run, max +5 ─────────────
    "00026": {
        COL_A1_TRIGGER:    "OnRun",
        COL_A1_TARGET:     "Self",
        COL_A1_FAIL_EVENT: None,
        COL_A1_COND:       None,
        COL_A1_COND_VAL:   None,
        COL_A1_COND2:      None,
        COL_A1_COND2_VAL:  None,
        COL_A1_EFFECT:     "StackStat",
        COL_A1_EFFECT_VAL: "RunBonus|Run|1|5",
        COL_A2_TRIGGER:    None,
        COL_A2_TARGET:     None,
        COL_A2_FAIL_EVENT: None,
        COL_A2_COND:       None,
        COL_A2_COND_VAL:   None,
        COL_A2_COND2:      None,
        COL_A2_COND2_VAL:  None,
        COL_A2_EFFECT:     None,
        COL_A2_EFFECT_VAL: None,
    },

    # ── Nash Thornton (QB): opponent discards on incomplete pass ─────────────
    "00048": {
        COL_A1_TRIGGER:    "OnPass",
        COL_A1_TARGET:     "None",
        COL_A1_FAIL_EVENT: None,
        COL_A1_COND:       "LastPassIncomplete",
        COL_A1_COND_VAL:   None,
        COL_A1_COND2:      None,
        COL_A1_COND2_VAL:  None,
        COL_A1_EFFECT:     "DiscardCard",
        COL_A1_EFFECT_VAL: "Opponent",
        COL_A2_TRIGGER:    None,
        COL_A2_TARGET:     None,
        COL_A2_FAIL_EVENT: None,
        COL_A2_COND:       None,
        COL_A2_COND_VAL:   None,
        COL_A2_COND2:      None,
        COL_A2_COND2_VAL:  None,
        COL_A2_EFFECT:     None,
        COL_A2_EFFECT_VAL: None,
    },

    # ── Damien Vaughn (RB): +1 run per consecutive run, max +3 ──────────────
    "00059": {
        COL_A1_TRIGGER:    "OnRun",
        COL_A1_TARGET:     "Self",
        COL_A1_FAIL_EVENT: None,
        COL_A1_COND:       None,
        COL_A1_COND_VAL:   None,
        COL_A1_COND2:      None,
        COL_A1_COND2_VAL:  None,
        COL_A1_EFFECT:     "StackStat",
        COL_A1_EFFECT_VAL: "RunBonus|Run|1|3",
        COL_A2_TRIGGER:    None,
        COL_A2_TARGET:     None,
        COL_A2_FAIL_EVENT: None,
        COL_A2_COND:       None,
        COL_A2_COND_VAL:   None,
        COL_A2_COND2:      None,
        COL_A2_COND2_VAL:  None,
        COL_A2_EFFECT:     None,
        COL_A2_EFFECT_VAL: None,
    },

    # ── Ibrahim Lewis (DB): +3 ShortCoverage and +3 DeepCoverage
    #                         fires on the 2nd+ consecutive pass ──────────────
    "00099": {
        COL_A1_TRIGGER:    "OnPass",
        COL_A1_TARGET:     "Self",
        COL_A1_FAIL_EVENT: None,
        COL_A1_COND:       "Consecutive",
        COL_A1_COND_VAL:   "ShortPass|>=2",
        COL_A1_COND2:      None,
        COL_A1_COND2_VAL:  None,
        COL_A1_EFFECT:     "AddStat",
        COL_A1_EFFECT_VAL: "ShortCoverage|3|1",
        # Ability 2: same trigger for deep pass streak
        COL_A2_TRIGGER:    "OnPass",
        COL_A2_TARGET:     "Self",
        COL_A2_FAIL_EVENT: None,
        COL_A2_COND:       "Consecutive",
        COL_A2_COND_VAL:   "LongPass|>=2",
        COL_A2_COND2:      None,
        COL_A2_COND2_VAL:  None,
        COL_A2_EFFECT:     "AddStat",
        COL_A2_EFFECT_VAL: "DeepCoverage|3|1",
    },

    # ── Xavier Morris (LB): +3 RunCoverage on 2nd+ consecutive run ──────────
    "00129": {
        COL_A1_TRIGGER:    "OnRun",
        COL_A1_TARGET:     "Self",
        COL_A1_FAIL_EVENT: None,
        COL_A1_COND:       "Consecutive",
        COL_A1_COND_VAL:   "Run|>=2",
        COL_A1_COND2:      None,
        COL_A1_COND2_VAL:  None,
        COL_A1_EFFECT:     "AddStat",
        COL_A1_EFFECT_VAL: "RunCoverage|3|1",
        COL_A2_TRIGGER:    None,
        COL_A2_TARGET:     None,
        COL_A2_FAIL_EVENT: None,
        COL_A2_COND:       None,
        COL_A2_COND_VAL:   None,
        COL_A2_COND2:      None,
        COL_A2_COND2_VAL:  None,
        COL_A2_EFFECT:     None,
        COL_A2_EFFECT_VAL: None,
    },
}

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: File not found: {XLSX_PATH}")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Build card_id → row mapping from column A
    id_to_row = {}
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        if cell.value is not None:
            id_to_row[str(cell.value).strip()] = cell.row

    updated = []
    missing = []
    for card_id, cols in PATCHES.items():
        if card_id not in id_to_row:
            missing.append(card_id)
            continue
        row_idx = id_to_row[card_id]
        for col_idx, value in cols.items():
            ws.cell(row=row_idx, column=col_idx).value = value
        updated.append(card_id)

    wb.save(XLSX_PATH)
    print(f"Saved {XLSX_PATH}")
    print(f"Updated ({len(updated)}): {', '.join(updated)}")
    if missing:
        print(f"NOT FOUND ({len(missing)}): {', '.join(missing)}")

if __name__ == "__main__":
    main()
