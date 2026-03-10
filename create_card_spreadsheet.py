"""
Creates First and Long - card_exporter.xlsx with:
  - Dropdown validation on all fixed-value columns
  - Color-coded column groups (stats / ability 1 / ability 2)
  - Frozen header row + column A
  - A "Reference" sheet with cheat codes for all cond/effect values
  - A "Probability" sheet with the slot outcome probability table

Run: python create_card_spreadsheet.py
Requires: pip install openpyxl
Output:  Assets/Resources/First and Long - card_exporter.xlsx
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = os.path.join(
    os.path.dirname(__file__),
    "Assets", "Resources", "First and Long - card_exporter.xlsx"
)

# ── Colours ──────────────────────────────────────────────────────────────────
C_HEADER      = "1A2E45"   # dark navy — header row
C_ID          = "E8EDF2"   # light grey — card_id / Name / pos / superstar / text
C_STATS       = "D6E8F5"   # light blue — stat columns
C_ABILITY1    = "D6F0DA"   # light green — ability slot 1
C_ABILITY2    = "FDEBD0"   # light orange — ability slot 2
C_REF_HEADER  = "2C3E50"   # reference sheet header
C_REF_ALT     = "F2F4F6"   # reference sheet alternating row

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font():
    return Font(bold=True, color="FFFFFF", size=10)

def cell_font(size=9):
    return Font(size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Column definitions ───────────────────────────────────────────────────────
# (header_text, width, bg_color, is_dropdown, dropdown_key)
ABILITY_COLS = [
    ("trigger",    12, None, True,  "trigger"),
    ("target",     10, None, True,  "target"),
    ("fail_event", 13, None, True,  "fail_event"),
    ("cond",       16, None, True,  "cond"),
    ("cond_val",   22, None, False, None),
    ("cond2",      16, None, True,  "cond"),
    ("cond2_val",  22, None, False, None),
    ("effect",     14, None, True,  "effect"),
    ("effect_val", 22, None, False, None),
]

BASE_COLS = [
    ("card_id",                  18, C_ID,    False, None),
    ("Name",                     20, C_ID,    False, None),
    ("pos",                       7, C_ID,    True,  "pos"),
    ("Superstar?",                9, C_ID,    True,  "superstar"),
    ("passive? (display text)",  30, C_ID,    False, None),
    ("Suit",                     10, C_ID,    True,  "suit"),
    ("Run Bonus",                10, C_STATS, False, None),
    ("Short Pass",               10, C_STATS, False, None),
    ("Deep Pass",                10, C_STATS, False, None),
    ("Run Coverage",             12, C_STATS, False, None),
    ("Short Coverage",           13, C_STATS, False, None),
    ("Deep Coverage",            13, C_STATS, False, None),
    ("Stamina",                   9, C_STATS, False, None),
    ("Grit",                      7, C_STATS, False, None),
]

def build_ability_cols(prefix, bg):
    return [(f"{prefix}_{c[0]}", c[1], bg, c[3], c[4]) for c in ABILITY_COLS]

ALL_COLS = (
    BASE_COLS
    + build_ability_cols("a1", C_ABILITY1)
    + build_ability_cols("a2", C_ABILITY2)
)

# ── Dropdown lists ───────────────────────────────────────────────────────────
DROPDOWNS = {
    "pos":        ["QB", "RB", "TE", "WR", "OL", "DL", "LB", "DB"],
    "superstar":  ["0", "1"],
    "suit":       ["Clubs", "Diamonds", "Hearts", "Spades"],
    "trigger":    ["OnPlay", "OnRun", "OnPass", "OnDraw", "OnDiscard",
                   "StartOfTurn", "EndOfTurn", "OnDriveStart", "OnIncomplete", "MANUAL"],
    "target":     ["Self", "AllBoard", "None"],
    "fail_event": ["", "TFL", "Sack", "Fumble", "QBFumble",
                   "Interception", "Incomplete", "BattedPass", "TippedPass"],
    "cond":       ["", "SlotIcon", "BoardCount", "CompareCounts", "Down",
                   "PlayType", "LastPlayType", "Consecutive", "FieldPos",
                   "HandCount", "StarCount", "PlaysLeft", "CoverageGuess",
                   "Random", "FirstSnap", "FirstPlay", "LastPassIncomplete",
                   "YardageGained", "StaminaCheck", "OncePerDrive",
                   "OncePerHalf", "OncePerGame", "PlayEnhancerPlayed",
                   "MANUAL"],
    "effect":     ["", "AddStat", "AddStatPerIcon", "ModifyStamina",
                   "DrawCard", "Knockout", "Charge", "DiscardCard",
                   "PreventLoss", "ShuffleFromDiscard", "AddSlotSymbol",
                   "ModifyGrit", "ReturnToHand", "Discover", "MANUAL"],
}

# ── Reference data ───────────────────────────────────────────────────────────
COND_REF = [
    # (shorthand, format, example, notes)
    ("SlotIcon",           "Icon|>=N",
     "Football|>=1",       "Icons: Football Helmet Star Wrench Wild"),
    ("BoardCount",         "Pos|Player|>=N",
     "OL|Self|>=3",        "Player: Self Opp Both  |  Pos: OL QB RB WR TE DL LB DB"),
    ("CompareCounts",      "CasterPos|OppPos|oper",
     "OL|DL|>",            "Caster position count vs opponent position count"),
    ("Down",               "N  or  >=N",
     "4",                  "Checks current down number"),
    ("PlayType",           "Run | ShortPass | LongPass",
     "Run",                "Checks current play type"),
    ("LastPlayType",       "Run | ShortPass | LongPass",
     "Run",                "Checks previous play type"),
    ("Consecutive",        "PlayType|>=N",
     "Run|>=2",            "N or more consecutive plays of that type"),
    ("FieldPos",           "Zone",
     "InRedZone",          "Zones: InRedZone  InOpponentTerritory  InOwnTerritory  GoalLine  BackedUp"),
    ("HandCount",          ">=N  or  <=N",
     ">=6",                "Cards in caster's hand"),
    ("StarCount",          "Pos|Player|>=N",
     "OL|Self|>=2",        "Counts superstar cards (isSuperstar=1) by position"),
    ("PlaysLeft",          "<=N",
     "<=5",                "Plays remaining in the half"),
    ("CoverageGuess",      "correct  or  wrong",
     "correct",            "Was the defense's play guess correct?"),
    ("Random",             "N  (0–100 integer %)",
     "25",                 "Pure chance — use only when no slot condition fits"),
    ("FirstSnap",          "(no params)",
     "",                   "First snap of the game"),
    ("FirstPlay",          "(no params)",
     "",                   "Card's first play after being placed on the field"),
    ("LastPassIncomplete", "(no params)",
     "",                   "Previous play was an incomplete pass"),
    ("YardageGained",      ">=N  or  <=N",
     ">=15",               "Checks yardage gained on this play"),
    ("StaminaCheck",       "Target|<=N  or  >=N",
     "Self|<=3",           "Target: Self  Opp  |  Checks stamina of target card"),
    ("OncePerDrive",       "(no params)",
     "",                   "Limits ability to fire once per drive"),
    ("OncePerHalf",        "(no params)",
     "",                   "Limits ability to fire once per half"),
    ("OncePerGame",        "(no params)",
     "",                   "Limits ability to fire once per game"),
    ("PlayEnhancerPlayed", "(no params)",
     "",                   "True if a play enhancer was played this turn"),
    ("MANUAL",             "—",
     "",                   "Skip — wire this ability in Unity Inspector"),
]

EFFECT_REF = [
    # (shorthand, format, example, notes)
    ("AddStat",         "StatName|amount|duration",
     "RunBonus|5|1",    "Stats: RunBonus  ShortPassBonus  DeepPassBonus\n"
                        "       RunCoverage  ShortCoverage  DeepCoverage\n"
                        "       Stamina  Grit\n"
                        "Duration 1 = this play only, 0 = permanent"),
    ("AddStatPerIcon",  "Icon|StatName|amountPerIcon",
     "Star|DeepPassBonus|3",
     "Multiplies icon count × amount.  E.g. 2 Stars × 3 = +6"),
    ("ModifyStamina",   "Target|amount",
     "Opp|-1",          "Target: Self  Team  Opp  |  Negative = remove stamina"),
    ("DrawCard",        "N",
     "1",               "Draw N cards for the caster's player"),
    ("Knockout",        "(no params)",
     "",                "Remove target card from the board"),
    ("Charge",          "Icon|threshold|StatName|amount",
     "Helmet|20|RunBonus|20",
     "Accumulate icon count over plays, fire when threshold met"),
    ("DiscardCard",     "Target|N",
     "Opp|1",           "Target: Self  Opp  |  Force discard of N cards"),
    ("PreventLoss",     "EventType",
     "Fumble",          "Negate a specific fail event (Fumble, Sack, etc.)"),
    ("ShuffleFromDiscard", "Pos",
     "OL",              "Shuffle a random card of Pos from discard into deck"),
    ("AddSlotSymbol",   "Icon|Reel",
     "Football|Left",   "Add an icon to a slot reel"),
    ("ModifyGrit",      "Target|amount",
     "Self|2",          "Target: Self  Team  Opp  |  Add/remove grit"),
    ("ReturnToHand",    "(no params)",
     "",                "Return caster from board to its owner's hand. Use with StartOfTurn + Down condition."),
    ("Discover",        "CardType|N",
     "OffensivePlayEnhancer|3",
     "Pull N random cards of CardType from deck, player picks 1 to keep, rest shuffled back"),
    ("MANUAL",          "—",
     "",                "Skip — wire this ability in Unity Inspector"),
]

PROB_REF = [
    ("Football ≥1",                    "80.5%", "Very common — nearly every spin"),
    ("Helmet ≥1",                      "70.7%", "Common"),
    ("Football ≥1  AND  Helmet ≥1",    "52.7%", "Near coin-flip"),
    ("Football ≥2",                    "37.5%", ""),
    ("Star ≥1",                        "33.0%", ""),
    ("Wrench ≥1",                      "23.4%", "Average QB incompletion threshold"),
    ("Helmet ≥2",                      "25.8%", ""),
    ("Star ≥1    AND  Football ≥1",    "22.9%", ""),
    ("Star ≥1    AND  Helmet ≥1",      "19.3%", ""),
    ("Wild ≥1",                        "12.5%", "Center reel only — lucky/special"),
    ("Football ≥3",                    " 7.0%", ""),
    ("Star ≥2",                        " 4.3%", "Rare explosive play"),
    ("Helmet ≥3",                      " 3.5%", ""),
    ("Wrench ≥2",                      " 1.6%", "Both outer reels — catastrophic"),
]

# ── Build Cards sheet ────────────────────────────────────────────────────────

def build_cards_sheet(ws):
    ws.title = "Cards"

    # Header row
    for col_idx, (header, width, bg, _, _) in enumerate(ALL_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font()
        cell.fill      = fill(C_HEADER)
        cell.alignment = center()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"   # freeze row 1 and column A

    # Data rows — colour bands + validation (rows 2..201)
    DATA_ROWS = 200
    dv_cache = {}  # key → DataValidation object

    for col_idx, (header, width, bg, is_dd, dd_key) in enumerate(ALL_COLS, start=1):
        col_letter = get_column_letter(col_idx)

        # Background fill for data cells
        if bg:
            for row in range(2, DATA_ROWS + 2):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill      = fill(bg)
                cell.font      = cell_font()
                cell.alignment = left()
                cell.border    = thin_border()

        # Dropdown validation
        if is_dd and dd_key:
            if dd_key not in dv_cache:
                opts = DROPDOWNS[dd_key]
                formula = '"' + ",".join(opts) + '"'
                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=True,
                    showDropDown=False,
                    showErrorMessage=True,
                    errorTitle="Invalid value",
                    error=f"Choose a value from the list (or leave blank).",
                )
                ws.add_data_validation(dv)
                dv_cache[dd_key] = dv
            dv_cache[dd_key].add(f"{col_letter}2:{col_letter}{DATA_ROWS + 1}")

    # Section divider lines — bold left border on first col of each group
    SECTION_STARTS = [1, 7, 15, 24]   # card_id, RunBonus, a1_trigger, a2_trigger
    for col_idx in SECTION_STARTS:
        thick = Side(style="medium", color="555555")
        for row in range(1, DATA_ROWS + 2):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=thick,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_COLS))}1"


# ── Build Reference sheet ────────────────────────────────────────────────────

def build_reference_sheet(ws):
    ws.title = "Reference"

    def hdr(row, col, text, width=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = fill(C_REF_HEADER)
        c.alignment = center()
        c.border    = thin_border()
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width

    def data(row, col, text, alt=False):
        c = ws.cell(row=row, column=col, value=text)
        # openpyxl auto-classifies any string starting with '=' as a formula.
        # Override to plain string so Excel doesn't try to evaluate it.
        if isinstance(text, str) and c.data_type == "f":
            c.data_type = "s"
        c.fill      = fill(C_REF_ALT) if alt else PatternFill()
        c.font      = cell_font(9)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border    = thin_border()

    # ── CONDITION reference ──────────────────────────────────────────────────
    r = 1
    ws.cell(row=r, column=1, value="CONDITION SHORTHANDS (a_cond / a_cond2)").font = Font(bold=True, size=12)
    r += 1
    for col, (text, w) in enumerate(
        [("Shorthand", 18), ("Format for cond_val", 26), ("Example", 22), ("Notes", 48)],
        start=1
    ):
        hdr(r, col, text, w)
    r += 1
    for i, (sh, fmt, ex, notes) in enumerate(COND_REF):
        alt = (i % 2 == 0)
        data(r, 1, sh,    alt); data(r, 2, fmt,   alt)
        data(r, 3, ex,    alt); data(r, 4, notes, alt)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1  # spacer

    # ── EFFECT reference ─────────────────────────────────────────────────────
    ws.cell(row=r, column=1, value="EFFECT SHORTHANDS (a_effect)").font = Font(bold=True, size=12)
    r += 1
    for col, (text, w) in enumerate(
        [("Shorthand", 18), ("Format for effect_val", 26), ("Example", 22), ("Notes", 48)],
        start=1
    ):
        hdr(r, col, text, w)
    r += 1
    for i, (sh, fmt, ex, notes) in enumerate(EFFECT_REF):
        alt = (i % 2 == 0)
        data(r, 1, sh,    alt); data(r, 2, fmt,   alt)
        data(r, 3, ex,    alt); data(r, 4, notes, alt)
        ws.row_dimensions[r].height = 40
        r += 1

    r += 1  # spacer

    # ── Operator quick reference ─────────────────────────────────────────────
    ws.cell(row=r, column=1, value="OPERATORS").font = Font(bold=True, size=12)
    r += 1
    hdr(r, 1, "Symbol", 10); hdr(r, 2, "Meaning", 20)
    r += 1
    for sym, meaning in [(">=", "greater than or equal"), (">", "greater than"),
                          ("<=", "less than or equal"),   ("<", "less than"),
                          ("==", "exactly equal"),         ("!=", "not equal")]:
        data(r, 1, sym); data(r, 2, meaning)
        r += 1

    r += 1

    # ── Position group quick reference ───────────────────────────────────────
    ws.cell(row=r, column=1, value="POSITION GROUPS").font = Font(bold=True, size=12)
    r += 1
    hdr(r, 1, "CSV pos", 10); hdr(r, 2, "Type", 16); hdr(r, 3, "Notes", 30)
    r += 1
    for pos, typ, note in [
        ("QB",  "Offensive", "Starts in sideline, not deck"),
        ("RB",  "Offensive", "RB_TE group"),
        ("TE",  "Offensive", "RB_TE group (same as RB)"),
        ("WR",  "Offensive", "WR group"),
        ("OL",  "Offensive", "OL group"),
        ("DL",  "Defensive", "DL group"),
        ("LB",  "Defensive", "LB group"),
        ("DB",  "Defensive", "DB group"),
    ]:
        data(r, 1, pos); data(r, 2, typ); data(r, 3, note)
        r += 1


# ── Build Probability sheet ──────────────────────────────────────────────────

def build_probability_sheet(ws):
    ws.title = "Probabilities"

    def hdr(row, col, text, width=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = fill(C_REF_HEADER)
        c.alignment = center()
        c.border    = thin_border()
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width

    r = 1
    ws.cell(row=r, column=1,
            value="Slot Machine Probabilities (middle row, 8×8×8 = 512 combos)").font = Font(bold=True, size=12)
    r += 1

    # Reel config
    ws.cell(row=r, column=1, value="Reel Layout:").font = Font(bold=True)
    r += 1
    for reel, comp in [
        ("Left  (R0)", "4 Football, 2 Helmet, 1 Star, 1 Wrench  — Wrench / no Wild"),
        ("Center(R1)", "3 Football, 3 Helmet, 1 Star, 1 Wild    — Wild / NO Wrench"),
        ("Right (R2)", "3 Football, 3 Helmet, 1 Star, 1 Wrench  — Wrench / no Wild"),
    ]:
        ws.cell(row=r, column=1, value=reel).font   = Font(bold=True, size=9)
        ws.cell(row=r, column=2, value=comp).font   = Font(size=9)
        r += 1

    r += 1
    hdr(r, 1, "Condition (cond + cond_val)", 36)
    hdr(r, 2, "Probability", 14)
    hdr(r, 3, "Notes", 40)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    r += 1

    from openpyxl.styles import PatternFill as PF
    for i, (cond, prob, notes) in enumerate(PROB_REF):
        alt = (i % 2 == 0)
        bg = PF("solid", fgColor=C_REF_ALT) if alt else PF()
        for col, val in [(1, cond), (2, prob), (3, notes)]:
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = bg
            c.font      = Font(size=9)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = thin_border()
        ws.row_dimensions[r].height = 16
        r += 1

    r += 2
    ws.cell(row=r, column=1,
            value="Design rule: Wrench = bad outcomes only (incompletions, fumbles, sacks, TFL).\n"
                  "Star = rare big plays.  Football = baseline success.  Wild = lucky/special (~12.5%).").font = Font(
        bold=True, size=9, color="8B0000")
    ws.row_dimensions[r].height = 32


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    cards_ws = wb.active
    build_cards_sheet(cards_ws)

    ref_ws = wb.create_sheet()
    build_reference_sheet(ref_ws)

    prob_ws = wb.create_sheet()
    build_probability_sheet(prob_ws)

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"  Cards sheet:        200 data rows, {len(ALL_COLS)} columns, dropdowns active")
    print(f"  Reference sheet:    condition/effect cheat codes")
    print(f"  Probabilities sheet: slot outcome table")


if __name__ == "__main__":
    main()
