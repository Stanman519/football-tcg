"""
Transpose all cards from raw_card_csv.csv (Part 1) and claud_raw_pt2_rough.txt (Part 2)
into the 'First and Long - card_exporter.xlsx' spreadsheet.

Fills: card_id, Name, pos, Superstar, passive text, Suit, stats, stamina, grit,
       and best-effort ability mapping (a1_* / a2_*). Cards that can't be auto-mapped
       get MANUAL triggers.

Run:  python transpose_cards_to_excel.py
"""

import csv
import os
import re
import openpyxl

BASE_DIR = os.path.dirname(__file__)
PART1_PATH = os.path.join(BASE_DIR, "Assets", "Resources", "raw_card_csv.csv")
PART2_PATH = os.path.join(BASE_DIR, "Assets", "Resources", "claud_raw_pt2_rough.txt")
EXCEL_PATH = os.path.join(BASE_DIR, "Assets", "Resources", "First and Long - card_exporter.xlsx")

OFFENSIVE_POS = {"QB", "RB", "WR", "TE", "OL"}
DEFENSIVE_POS = {"DL", "LB", "DB"}

# Excel column indices (1-based) after adding Suit column
COL = {
    "card_id": 1, "Name": 2, "pos": 3, "Superstar": 4, "passive": 5, "Suit": 6,
    "RunBonus": 7, "ShortPass": 8, "DeepPass": 9,
    "RunCov": 10, "ShortCov": 11, "DeepCov": 12,
    "Stamina": 13, "Grit": 14,
    # Ability 1
    "a1_trigger": 15, "a1_target": 16, "a1_fail_event": 17,
    "a1_cond": 18, "a1_cond_val": 19, "a1_cond2": 20, "a1_cond2_val": 21,
    "a1_effect": 22, "a1_effect_val": 23,
    # Ability 2
    "a2_trigger": 24, "a2_target": 25, "a2_fail_event": 26,
    "a2_cond": 27, "a2_cond_val": 28, "a2_cond2": 29, "a2_cond2_val": 30,
    "a2_effect": 31, "a2_effect_val": 32,
}

# ── Report collector ─────────────────────────────────────────────────────────
report_manual = []   # (card_id, name, reason)
report_new_mechanics = []  # (card_id, name, mechanic description)
report_questions = []  # (card_id, name, question)


def safe_int(val, default=0):
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def format_card_id(num):
    return f"{num:05d}"


# ── Ability mapping engine ───────────────────────────────────────────────────
# Returns a dict with keys like a1_trigger, a1_cond, etc.
# If it returns None, the card needs MANUAL wiring.

def parse_tags(tags_str):
    """Split tags on commas, spaces, AND underscores — but preserve known compound tokens."""
    COMPOUND_TAGS = {"LOSS_PREV", "FIRST_SNAP", "RUN_FUMBLE"}
    raw_tokens = [t.strip().upper() for t in re.split(r'[,\s]+', tags_str) if t.strip()]
    result = set()
    for token in raw_tokens:
        if not token:
            continue
        if "_" not in token or token in COMPOUND_TAGS:
            result.add(token)
        else:
            # Split on underscore — each segment is an individual tag
            for part in token.split("_"):
                if part:
                    # Recombine known compounds (LOSS + PREV → LOSS_PREV)
                    result.add(part)
            # Also keep the full token for specific compound checks
            result.add(token)
    return result


def map_ability(card_id, name, pos, passive, tags_str, ability_opt=""):
    """
    Attempt to map a card's passive/tags into structured ability columns.
    Returns (a1_dict, a2_dict, manual_reason) where dicts have ability field values.
    """
    tags = parse_tags(tags_str)
    p = passive.lower().strip() if passive else ""
    ao = ability_opt.lower().strip() if ability_opt else ""

    a1 = {}
    a2 = {}
    manual_reason = None

    # Helper to determine trigger from context
    def guess_trigger():
        if "RUN" in tags and "SP" not in tags and "DP" not in tags:
            return "OnRun"
        if ("SP" in tags or "DP" in tags) and "RUN" not in tags:
            return "OnPass"
        return "OnPlay"

    # Helper to determine the stat being modified
    def guess_stat(pos):
        if pos in OFFENSIVE_POS:
            if "RUN" in tags and "SP" not in tags and "DP" not in tags:
                return "RunBonus"
            if "SP" in tags and "DP" not in tags:
                return "ShortPassBonus"
            if "DP" in tags and "SP" not in tags:
                return "DeepPassBonus"
        else:  # defensive
            if "RUN" in tags and "SP" not in tags and "DP" not in tags:
                return "RunCoverage"
            if "SP" in tags and "DP" not in tags:
                return "ShortCoverage"
            if "DP" in tags and "SP" not in tags:
                return "DeepCoverage"
        return None

    # ── CHARGE cards → always MANUAL ────────────────────────────────────────
    if "CHARGE" in tags:
        manual_reason = f"Charge mechanic: {passive}"
        a1 = {"a1_trigger": "MANUAL", "a1_effect": "Charge"}
        return a1, a2, manual_reason

    # ── DISCARD-as-cost cards → always MANUAL ───────────────────────────────
    if "DISCARD" in tags and ("discard" in p and ("before" in p or "you may discard" in p or "cost" in p)):
        manual_reason = f"Discard-as-cost: {passive}"
        a1 = {"a1_trigger": "MANUAL", "a1_effect": "DiscardCard"}
        return a1, a2, manual_reason

    # ── LOSS_PREV cards → MANUAL (prevent loss events) ──────────────────────
    if "LOSS_PREV" in tags:
        manual_reason = f"Prevent-loss mechanic: {passive}"
        a1 = {"a1_trigger": "MANUAL", "a1_effect": "PreventLoss"}
        return a1, a2, manual_reason

    # ── SLOTMANIP cards → MANUAL ────────────────────────────────────────────
    if "SLOTMANIP" in tags and ("add" in p and "slot" in p):
        manual_reason = f"Slot manipulation: {passive}"
        a1 = {"a1_trigger": "MANUAL", "a1_effect": "AddSlotSymbol"}
        return a1, a2, manual_reason

    # ── KNOCKOUT (remove stamina from opponent) ─────────────────────────────
    if "KNOCKOUT" in tags:
        manual_reason = f"Knockout/stamina removal: {passive}"
        a1 = {"a1_trigger": "MANUAL", "a1_effect": "Knockout"}
        return a1, a2, manual_reason

    # ── SEQ cards with stacking/resetting counters → MANUAL ─────────────────
    if "SEQ" in tags and ("consecutive" in p or "resets" in p or "stacking" in p or "every 2 plays" in p):
        manual_reason = f"Sequential/stacking mechanic: {passive}"
        a1 = {"a1_trigger": "MANUAL"}
        return a1, a2, manual_reason

    # ── DRAW cards (draw on condition) ──────────────────────────────────────
    if "DRAW" in tags and ("draw" in p):
        # Try to figure out the condition
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "None"
        a1["a1_effect"] = "DrawCard"
        a1["a1_effect_val"] = "1"

        if "skip" in p and "draw" in p:
            manual_reason = f"Skip-draw mechanic: {passive}"
            a1["a1_trigger"] = "MANUAL"
            a1["a1_effect"] = "SkipDraw"
            return a1, a2, manual_reason

        # Condition parsing for draw cards
        if "15 or more yards" in p or "15+" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = ">=15"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "7 or fewer" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = "<=7"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "8 or fewer" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = "<=8"
            if "1st or 2nd" in p or "1st/2nd" in p:
                a1["a1_cond2"] = "Down"
                a1["a1_cond2_val"] = "<=2"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "incomplete" in p:
            a1["a1_cond"] = "LastPassIncomplete"
        elif "last play was" in p and "run" in p:
            a1["a1_cond"] = "LastPlayType"
            a1["a1_cond_val"] = "Run"
        elif "short pass" in p and "completed" in p and "last" in p:
            a1["a1_cond"] = "LastPlayType"
            a1["a1_cond_val"] = "ShortPass"
            a1["a1_trigger"] = "OnPlay"

        return a1, a2, manual_reason

    # ── TURNOVER cards → MANUAL ─────────────────────────────────────────────
    if "TURNOVER" in tags or "FUMBLE" in tags:
        manual_reason = f"Turnover/fumble mechanic: {passive}"
        a1 = {"a1_trigger": "MANUAL"}
        if "fumble" in p:
            a1["a1_fail_event"] = "Fumble"
            a1["a1_cond"] = "Random"
            # Try to extract percentage
            pct_match = re.search(r'(\d+)\s*percent', p)
            if pct_match:
                a1["a1_cond_val"] = pct_match.group(1)
        return a1, a2, manual_reason

    # ── GRIT modification cards ─────────────────────────────────────────────
    if "GRIT" in tags and ("grit" in p):
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_effect"] = "ModifyGrit"

        # Try to extract grit amount
        grit_match = re.search(r'add\s*(\d+)\s*(?:to\s*)?(?:your\s*)?(?:grit|defensive)', p)
        if grit_match:
            a1["a1_effect_val"] = f"Self|{grit_match.group(1)}"
        else:
            grit_match = re.search(r'(\d+)\s*grit', p)
            if grit_match:
                a1["a1_effect_val"] = f"Self|{grit_match.group(1)}"

        # Check for conditions
        if "4th down" in p:
            a1["a1_cond"] = "Down"
            a1["a1_cond_val"] = "4"
        elif "1st down" in p:
            a1["a1_cond"] = "Down"
            a1["a1_cond_val"] = "1"
        elif "incomplete" in p:
            a1["a1_cond"] = "LastPassIncomplete"
        elif "grit is" in p and ("higher" in p or "tied" in p or "greater" in p):
            manual_reason = f"Grit comparison condition: {passive}"
        elif "grit is lower" in p:
            manual_reason = f"Grit comparison condition: {passive}"

        if "STARQTY" in tags:
            manual_reason = f"Grit + star count condition: {passive}"

        return a1, a2, manual_reason

    # ── HANDQTY cards (hand count conditions) ───────────────────────────────
    if "HANDQTY" in tags:
        if "discard" in p:
            manual_reason = f"Discard-as-cost + hand count: {passive}"
            a1 = {"a1_trigger": "MANUAL"}
            return a1, a2, manual_reason
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "HandCount"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        # Try to extract hand count threshold
        hand_match = re.search(r'(\d+)\s*or\s*more\s*cards\s*in\s*hand', p)
        if hand_match:
            a1["a1_cond_val"] = f">={hand_match.group(1)}"
        elif "empty" in p:
            a1["a1_cond_val"] = "<=0"
        # Try to extract bonus amount
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── GUESS (coverage guess) cards ────────────────────────────────────────
    if "GUESS" in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "CoverageGuess"
        if "wrong" in p or "incorrect" in p:
            a1["a1_cond_val"] = "wrong"
        else:
            a1["a1_cond_val"] = "correct"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        elif stat:
            # Try "reduce ... by N"
            reduce_match = re.search(r'reduce\s*\w*\s*(?:run|pass|short|deep)\s*(?:pass\s*)?by\s*(?:an\s*extra\s*)?(\d+)', p)
            if reduce_match:
                a1["a1_effect_val"] = f"{stat}|{reduce_match.group(1)}|1"

        if "each helmet" in p or "per helmet" in p or "for each helmet" in p:
            a1["a1_effect"] = "AddStatPerIcon"
            icon_match = re.search(r'(\d+)\s*yard', p)
            if icon_match and stat:
                a1["a1_effect_val"] = f"Helmet|{stat}|{icon_match.group(1)}"

        return a1, a2, manual_reason

    # ── DOWN (down-based bonuses) ───────────────────────────────────────────
    if "DOWN" in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "Down"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)

        if "4th down" in p or "4th downs" in p:
            a1["a1_cond_val"] = "4"
        elif "3rd down" in p or "3rd downs" in p:
            a1["a1_cond_val"] = "3"
        elif "1st down" in p or "first down" in p or "first downs" in p:
            a1["a1_cond_val"] = "1"
        elif "1st or 2nd" in p or "2nd down" in p:
            a1["a1_cond_val"] = "<=2"
        elif "2nd down" in p:
            a1["a1_cond_val"] = "2"

        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"

        # Special: "sit out every 3rd down"
        if "sit out" in p:
            a1["a1_effect"] = "SitOut"
            a1["a1_effect_val"] = a1.get("a1_cond_val", "3")
            manual_reason = f"SitOut mechanic (not yet coded): {passive}"

        # Special: first snap + 4th down compound
        if "first snap" in p:
            a1["a1_cond2"] = "FirstSnap"
            a1["a1_cond2_val"] = ""

        # "add 1 stamina" type
        if "stamina" in p:
            a1["a1_effect"] = "ModifyStamina"
            stam_match = re.search(r'(\d+)\s*stamina', p)
            if stam_match:
                a1["a1_effect_val"] = f"Team|{stam_match.group(1)}"
            a1["a1_target"] = "AllBoard"

        return a1, a2, manual_reason

    # ── SLOTDRAW (slot icon conditions) ─────────────────────────────────────
    if "SLOTDRAW" in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "SlotIcon"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)

        # Try to identify slot icon type and threshold
        if "star" in p:
            icon = "Star"
        elif "helmet" in p:
            icon = "Helmet"
        elif "football" in p:
            icon = "Football"
        elif "wrench" in p:
            icon = "Wrench"
        elif "wild" in p:
            icon = "Wild"
        else:
            icon = "Football"  # default

        # Check for count
        count_match = re.search(r'(\d+)\s*' + icon.lower(), p)
        if count_match:
            a1["a1_cond_val"] = f"{icon}|>={count_match.group(1)}"
        elif "at least 1" in p or "1+" in p:
            a1["a1_cond_val"] = f"{icon}|>=1"
        elif "2 or more" in p or "2+" in p:
            a1["a1_cond_val"] = f"{icon}|>=2"
        else:
            a1["a1_cond_val"] = f"{icon}|>=1"

        # Percentage-based (use Random instead)
        pct_match = re.search(r'(\d+)\s*(?:percent|%)', p)
        if pct_match:
            a1["a1_cond"] = "Random"
            a1["a1_cond_val"] = pct_match.group(1)

        # "per star" / "per icon" → AddStatPerIcon
        if "per star" in p or "each star" in p or "every star" in p or "for each star" in p:
            a1["a1_effect"] = "AddStatPerIcon"
            amt_match = re.search(r'\+(\d+)\s*(?:yard|bonus)', p)
            if amt_match and stat:
                a1["a1_effect_val"] = f"Star|{stat}|{amt_match.group(1)}"
            elif stat:
                a1["a1_effect_val"] = f"Star|{stat}|1"
            return a1, a2, manual_reason

        # Wrench = penalty (negative stat)
        if "wrench" in p and ("decrease" in p or "reduce" in p or "lose" in p or "subtract" in p):
            bonus_match = re.search(r'(\d+)', p)
            if bonus_match and stat:
                a1["a1_effect_val"] = f"{stat}|-{bonus_match.group(1)}|1"
            return a1, a2, manual_reason

        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"

        # Cards with both positive and negative ("add X if star else subtract Y")
        if "else" in p and "subtract" in p:
            manual_reason = f"Conditional positive/negative based on slot: {passive}"

        # 25% = use Random(25) since no single slot condition maps to 25%
        if "25%" in p:
            a1["a1_cond"] = "Random"
            a1["a1_cond_val"] = "25"

        return a1, a2, manual_reason

    # ── BOOST (board position count bonuses) ────────────────────────────────
    if "BOOST" in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)

        # "if you have X or more [POS]" → BoardCount
        bc_match = re.search(r'(\d+)\s*or\s*more\s*(WR|TE|OL|RB|QB|DL|LB|DB)', p, re.IGNORECASE)
        if bc_match:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = f"{bc_match.group(2).upper()}|Self|>={bc_match.group(1)}"
        elif "another wr" in p or "other wr" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "WR|Self|>=2"
        elif "exactly 1 other wr" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "WR|Self|==2"  # self + 1 other = 2
        elif "te on the field" in p or "te on field" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "TE|Self|>=1"
        elif "a te" in p and "not on field" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "TE|Opp|==0"  # opponent's TE... wait, this is about offense not having a TE
            # Actually for defensive cards checking if offense doesn't have TE
            a1["a1_cond_val"] = "TE|Opp|==0"

        # "add +N to other [POS]" → AllBoard target
        if "other" in p and ("o line" in p or "ol" in p):
            a1["a1_target"] = "AllBoard"

        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"

        # STARQTY + BOOST → star count condition
        if "STARQTY" in tags:
            if "star ol" in p or "star o line" in p:
                a1["a1_cond"] = "StarCount"
                sc_match = re.search(r'(\d+)\s*or\s*more\s*star', p)
                if sc_match:
                    a1["a1_cond_val"] = f"OL|Self|>={sc_match.group(1)}"
                else:
                    a1["a1_cond_val"] = "OL|Self|>=1"
            elif "star dl" in p or "star d line" in p:
                a1["a1_cond"] = "StarCount"
                sc_match = re.search(r'(\d+)\s*or\s*more\s*star', p)
                if sc_match:
                    a1["a1_cond_val"] = f"DL|Self|>={sc_match.group(1)}"
            elif "qb is a superstar" in p:
                a1["a1_cond"] = "StarCount"
                a1["a1_cond_val"] = "QB|Self|>=1"

        return a1, a2, manual_reason

    # ── OLPOS / DEFPOS (position count comparisons) ─────────────────────────
    if "OLPOS" in tags or "DEFPOS" in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)

        if "more dl than" in p or "more d line than" in p:
            a1["a1_cond"] = "CompareCounts"
            a1["a1_cond_val"] = "DL|OL|>"
        elif "more ol than" in p or "more o line than" in p or "ol > dl" in ao:
            a1["a1_cond"] = "CompareCounts"
            a1["a1_cond_val"] = "OL|DL|>"
        elif "ol >" in p and "total def" in p:
            a1["a1_cond"] = "CompareCounts"
            a1["a1_cond_val"] = "OL|DEF|>"
        elif "3 or more ol" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "OL|Self|>=3"
        elif "2 or more ol" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "OL|Self|>=2"
        elif "3 or more db" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "DB|Self|>=3"
        elif "3 or more lb" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "LB|Self|>=3"
        elif "1 or fewer ol" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "OL|Opp|<=1"
        elif "1 or fewer lb" in p:
            a1["a1_cond"] = "BoardCount"
            a1["a1_cond_val"] = "LB|Opp|<=1"
        elif "2 or fewer" in p and "star" in p:
            a1["a1_cond"] = "StarCount"
            a1["a1_cond_val"] = "ALL|Opp|<=2"
        elif "no dl star" in p or "no d line star" in p:
            a1["a1_cond"] = "StarCount"
            a1["a1_cond_val"] = "DL|Opp|==0"
        elif "each ol" in p or "every ol" in p or "for each o line" in p:
            # Per OL bonus
            a1["a1_effect"] = "MANUAL"
            manual_reason = f"Per-OL-count bonus: {passive}"
        elif "rb" in p and "stamina" in p:
            manual_reason = f"RB stamina check: {passive}"
            a1["a1_trigger"] = "MANUAL"

        # STARQTY overlaps
        if "STARQTY" in tags and "star" in p:
            if "more dl stars" in p or "more d line stars" in p:
                a1["a1_cond"] = "CompareCounts"
                a1["a1_cond_val"] = "DL*|OL*|>"
                manual_reason = f"Star-count comparison (needs CompareCounts star variant): {passive}"
            elif "total star" in p and "more" in p:
                manual_reason = f"Total star comparison: {passive}"

        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"

        if "reduce" in p:
            reduce_match = re.search(r'reduce\b.{0,40}?\bby\s+(\d+)', p)
            if reduce_match and stat:
                a1["a1_effect_val"] = f"{stat}|{reduce_match.group(1)}|1"

        return a1, a2, manual_reason

    # ── STARQTY (star count) without BOOST or DEFPOS ────────────────────────
    if "STARQTY" in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)

        if "star ol" in p or "star o line" in p:
            a1["a1_cond"] = "StarCount"
            sc_match = re.search(r'(\d+)\s*or\s*more\s*star', p)
            if sc_match:
                a1["a1_cond_val"] = f"OL|Self|>={sc_match.group(1)}"
            else:
                a1["a1_cond_val"] = "OL|Self|>=1"
        elif "star dl" in p or "star d line" in p:
            a1["a1_cond"] = "StarCount"
            sc_match = re.search(r'(\d+)\s*or\s*more\s*star', p)
            if sc_match:
                a1["a1_cond_val"] = f"DL|Self|>={sc_match.group(1)}"
        elif "star player" in p and ("1" in p or "at least" in p):
            a1["a1_cond"] = "StarCount"
            a1["a1_cond_val"] = "ALL|Self|>=1"
        elif "total star" in p and "2 more" in p:
            a1["a1_cond"] = "StarCount"
            manual_reason = f"Star count differential: {passive}"

        if "first snap" in p:
            a1["a1_cond"] = "FirstSnap"
            a1["a1_cond2"] = "StarCount"
            a1["a1_cond2_val"] = "OL|Self|>=1"  # generic

        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"

        return a1, a2, manual_reason

    # ── SEQ (sequential / last-play conditions) ─────────────────────────────
    if "SEQ" in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)

        if "preceded by a run" in p or "previous play was a run" in p or "last play was a run" in p:
            a1["a1_cond"] = "LastPlayType"
            a1["a1_cond_val"] = "Run"
        elif "preceded by a short pass" in p or "last play was a short pass" in p or "short pass last play" in p:
            a1["a1_cond"] = "LastPlayType"
            a1["a1_cond_val"] = "ShortPass"
        elif "completed a short pass last play" in p or "short pass was completed last play" in p:
            a1["a1_cond"] = "LastPlayType"
            a1["a1_cond_val"] = "ShortPass"
        elif "run of 10 or more" in p or "run gained 10" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = ">=10"
            a1["a1_cond2"] = "LastPlayType"
            a1["a1_cond2_val"] = "Run"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "15 or more yards last play" in p or "gained 15 or more yards" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = ">=15"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "lost yardage" in p or "previous play lost" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = "<0"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "incomplete" in p:
            a1["a1_cond"] = "LastPassIncomplete"
        elif "correct pass coverage" in p or "correct coverage" in p:
            a1["a1_cond"] = "CoverageGuess"
            a1["a1_cond_val"] = "correct"
        elif "consecutive" in p and "run" in p:
            a1["a1_cond"] = "Consecutive"
            a1["a1_cond_val"] = "Run|>=2"
            manual_reason = f"Consecutive mechanic with reset: {passive}"
        elif "consecutive" in p and "pass" in p:
            a1["a1_cond"] = "Consecutive"
            a1["a1_cond_val"] = "ShortPass|>=2"
            manual_reason = f"Consecutive mechanic: {passive}"

        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"

        if "reduce" in p:
            reduce_match = re.search(r'reduce\b.{0,40}?\bby\s+(\d+)', p)
            if reduce_match and stat:
                a1["a1_effect_val"] = f"{stat}|{reduce_match.group(1)}|1"

        return a1, a2, manual_reason

    # ── FIELDPOS (red zone etc) ─────────────────────────────────────────────
    if "red zone" in p or "inside the" in p:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "FieldPos"
        a1["a1_cond_val"] = "InRedZone"
        if "inside the 5" in p:
            a1["a1_cond_val"] = "GoalLine"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"

        # Special: "always score inside 5"
        if "always score" in p:
            manual_reason = f"Auto-touchdown mechanic: {passive}"
            a1["a1_effect"] = "MANUAL"

        return a1, a2, manual_reason

    # ── Simple stat add with no special tag ─────────────────────────────────
    # Cards like "+X stat when pass coverage" (GUESS tagged but no special tags)
    if "pass coverage" in p and ("GUESS" not in tags):
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "CoverageGuess"
        a1["a1_cond_val"] = "wrong"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── Wrenches decrease bonus ─────────────────────────────────────────────
    if "wrench" in p and ("decrease" in p or "reduce" in p):
        a1["a1_trigger"] = "OnPlay"
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "SlotIcon"
        a1["a1_cond_val"] = "Wrench|>=1"
        a1["a1_effect"] = "AddStatPerIcon"
        stat = guess_stat(pos) or "ShortPassBonus"
        a1["a1_effect_val"] = f"Wrench|{stat}|-1"
        return a1, a2, manual_reason

    # ── "does not count toward" mechanic → MANUAL ───────────────────────────
    if "does not count" in p:
        manual_reason = f"Roster limit bypass: {passive}"
        a1 = {"a1_trigger": "MANUAL"}
        return a1, a2, manual_reason

    # ── "+X if Y condition" with no matching tags ───────────────────────────
    if "slot trigger" in p or "slot" in p and "incomplete" in p:
        # Checkdown-type
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "SlotIcon"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        # percentage
        pct_match = re.search(r'(\d+)\s*percent', p)
        if pct_match:
            a1["a1_cond"] = "Random"
            a1["a1_cond_val"] = pct_match.group(1)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── DISCARD tag without cost pattern ─────────────────────────────────────
    if "DISCARD" in tags:
        manual_reason = f"Discard mechanic: {passive}"
        a1 = {"a1_trigger": "MANUAL", "a1_effect": "DiscardCard"}
        return a1, a2, manual_reason

    # ══════════════════════════════════════════════════════════════════════════
    # PASSIVE-TEXT-BASED PATTERN MATCHING (catches cards with missing/bad tags)
    # ══════════════════════════════════════════════════════════════════════════

    # ── Stamina check: "if X or less stamina" ───────────────────────────────
    stam_match = re.search(r'(\d+)\s*or\s*(?:less|fewer)\s*stamina', p)
    if stam_match:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "StaminaCheck"
        a1["a1_cond_val"] = f"Self|<={stam_match.group(1)}"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        manual_reason = f"StaminaCheck condition (not yet coded): {passive}"
        return a1, a2, manual_reason

    # ── Hand count: "X or more cards in hand" ───────────────────────────────
    hand_match = re.search(r'(\d+)\s*or\s*more\s*cards\s*in\s*hand', p)
    if hand_match:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "HandCount"
        a1["a1_cond_val"] = f">={hand_match.group(1)}"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    if "hand is empty" in p or "no cards in hand" in p:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "HandCount"
        a1["a1_cond_val"] = "<=0"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        if "else" in p:
            manual_reason = f"Conditional hand-empty: {passive}"
        return a1, a2, manual_reason

    # ── Board count from passive text (no BOOST/DEFPOS tag) ─────────────────
    bc_text = re.search(r'(\d+)\s*or\s*more\s*(OL|WR|TE|RB|QB|DL|LB|DB)s?\s*(?:in play|on the field|on field)', p, re.IGNORECASE)
    if bc_text:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "BoardCount"
        a1["a1_cond_val"] = f"{bc_text.group(2).upper()}|Self|>={bc_text.group(1)}"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── "if TE on the field" / "if TE is not on field" ──────────────────────
    if ("te on the field" in p or "te on field" in p or "have a te" in p) and "BOOST" not in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "BoardCount"
        a1["a1_cond_val"] = "TE|Self|>=1"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    if "te is not on field" in p or "te not on field" in p or "no te" in p:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "BoardCount"
        a1["a1_cond_val"] = "TE|Opp|==0"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── "X or fewer LB/DL/DB" (defense has few of a position) ───────────────
    few_pos = re.search(r'(\d+)\s*or\s*fewer\s*(LB|DL|DB|OL|WR|TE)s?', p, re.IGNORECASE)
    if few_pos:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "BoardCount"
        a1["a1_cond_val"] = f"{few_pos.group(2).upper()}|Opp|<={few_pos.group(1)}"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── StarCount from passive: "no DL star" / "X or fewer stars" ───────────
    if "no dl star" in p or "no d line star" in p or "no dl star" in p:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "StarCount"
        a1["a1_cond_val"] = "DL|Opp|==0"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    star_few = re.search(r'(\d+)\s*or\s*fewer\s*stars?', p)
    if star_few:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "StarCount"
        a1["a1_cond_val"] = f"ALL|Opp|<={star_few.group(1)}"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── CoverageGuess from passive text ─────────────────────────────────────
    if "guess" in p and ("incorrect" in p or "wrong" in p):
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "CoverageGuess"
        a1["a1_cond_val"] = "wrong"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    if "pass coverage" in p and "GUESS" not in tags:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "CoverageGuess"
        a1["a1_cond_val"] = "wrong"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── CompareCounts from passive text ─────────────────────────────────────
    cmp_match = re.search(r'more\s*(OL|DL|LB|DB|WR)\s*than.{0,20}?(OL|DL|LB|DB|WR|defensive)', p, re.IGNORECASE)
    if cmp_match:
        caster = cmp_match.group(1).upper()
        target = cmp_match.group(2).upper()
        if target == "DEFENSIVE":
            target = "DEF"
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "CompareCounts"
        a1["a1_cond_val"] = f"{caster}|{target}|>"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── LastPlayType from passive text ──────────────────────────────────────
    if "last play" in p or "previous play" in p or "preceded by" in p:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        if "run" in p and ("short pass" not in p and "pass" not in p):
            a1["a1_cond"] = "LastPlayType"
            a1["a1_cond_val"] = "Run"
        elif "short pass" in p:
            a1["a1_cond"] = "LastPlayType"
            a1["a1_cond_val"] = "ShortPass"
        elif "lost yardage" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = "<0"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "15 or more" in p or "15+" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = ">=15"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        else:
            manual_reason = f"Could not parse last-play condition: {passive}"
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── PlayEnhancerPlayed from passive text ────────────────────────────────
    if "play enhancer" in p:
        a1["a1_trigger"] = "OnPlay"
        a1["a1_target"] = "Self"
        a1["a1_cond"] = "PlayEnhancerPlayed"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        manual_reason = f"PlayEnhancerPlayed condition (not yet coded): {passive}"
        return a1, a2, manual_reason

    # ── Draw card from passive text ─────────────────────────────────────────
    if "draw" in p and ("card" in p or "1" in p):
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "None"
        a1["a1_effect"] = "DrawCard"
        a1["a1_effect_val"] = "1"
        # Try to determine condition
        if "15 or more" in p or "15+" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = ">=15"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "8 or fewer" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = "<=8"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "7 or fewer" in p:
            a1["a1_cond"] = "YardageGained"
            a1["a1_cond_val"] = "<=7"
            manual_reason = f"YardageGained condition (not yet coded): {passive}"
        elif "incomplete" in p:
            a1["a1_cond"] = "LastPassIncomplete"
        elif "short pass" in p and "completed" in p:
            a1["a1_cond"] = "LastPlayType"
            a1["a1_cond_val"] = "ShortPass"
        return a1, a2, manual_reason

    # ── Skip draw mechanic ──────────────────────────────────────────────────
    if "skip" in p and "draw" in p:
        stat = guess_stat(pos)
        bonus_match = re.search(r'\+(\d+)', p)
        a1["a1_trigger"] = "MANUAL"
        a1["a1_effect"] = "SkipDraw"
        if bonus_match and stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}"
        manual_reason = f"SkipDraw mechanic (not yet coded): {passive}"
        return a1, a2, manual_reason

    # ── Fallback: try to extract any "+N" stat bonus ────────────────────────
    bonus_match = re.search(r'\+(\d+)', p)
    if bonus_match:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        if stat:
            a1["a1_effect_val"] = f"{stat}|{bonus_match.group(1)}|1"
        # Try basic condition from passive text
        if "4th down" in p:
            a1["a1_cond"] = "Down"
            a1["a1_cond_val"] = "4"
        elif "3rd down" in p:
            a1["a1_cond"] = "Down"
            a1["a1_cond_val"] = "3"
        elif "1st down" in p or "first down" in p:
            a1["a1_cond"] = "Down"
            a1["a1_cond_val"] = "1"
        elif "first snap" in p:
            a1["a1_cond"] = "FirstSnap"
        elif "grit" in p and ("higher" in p or "tied" in p):
            a1["a1_cond"] = "MANUAL"
            manual_reason = f"Grit comparison: {passive}"
        elif "consecutive" in p:
            a1["a1_cond"] = "Consecutive"
            if "run" in p:
                a1["a1_cond_val"] = "Run|>=2"
            elif "pass" in p:
                a1["a1_cond_val"] = "ShortPass|>=2"
            manual_reason = f"Consecutive mechanic: {passive}"
        elif "every ol" in p or "each ol" in p or "for every ol" in p or "for each o line" in p:
            manual_reason = f"Per-position-count bonus: {passive}"
        else:
            manual_reason = f"Could not determine condition: {passive}"
        return a1, a2, manual_reason

    # ── "reduce" patterns ────────────────────────────────────────────────────
    reduce_match = re.search(r'reduce\b.{0,40}?\bby\s+(\d+)', p)
    if reduce_match:
        a1["a1_trigger"] = guess_trigger()
        a1["a1_target"] = "Self"
        a1["a1_effect"] = "AddStat"
        stat = guess_stat(pos)
        if stat:
            a1["a1_effect_val"] = f"{stat}|{reduce_match.group(1)}|1"
        return a1, a2, manual_reason

    # ── No passive text at all ──────────────────────────────────────────────
    if not p:
        return a1, a2, None  # vanilla card, no ability

    # ── Catch-all → MANUAL ──────────────────────────────────────────────────
    manual_reason = f"Could not auto-map: {passive}"
    a1 = {"a1_trigger": "MANUAL"}
    return a1, a2, manual_reason


# ── CSV parsing ──────────────────────────────────────────────────────────────

def read_part1():
    """Read raw_card_csv.csv (Part 1: OL, QB, RB).
    Data rows have 18 columns (header has 17) — extra empty col between Parameters and Tags.
    Real layout: [0-12]=basic, [13]+[14]=parameters (one is always empty), [15]=Tags, [16]=ability_optimized, [17]=Suit.
    """
    cards = []
    with open(PART1_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if len(row) < 8 or not row[0].strip():
                continue
            # Merge the two parameter columns (one is always empty)
            params = safe_str(row[13]) if len(row) > 13 else ""
            if not params and len(row) > 14:
                params = safe_str(row[14])
            cards.append({
                "name": safe_str(row[0]),
                "pos": safe_str(row[1]).upper(),
                "superstar": safe_str(row[2]),
                "passive": safe_str(row[3]),
                "run_mod": safe_int(row[4]),
                "short_mod": safe_int(row[5]),
                "deep_mod": safe_int(row[6]),
                "stamina": safe_int(row[7]),
                "grit": safe_int(row[8]) if len(row) > 8 else 0,
                "target": safe_str(row[9]) if len(row) > 9 else "",
                "bonus_amount": safe_str(row[10]) if len(row) > 10 else "",
                "affected_stat": safe_str(row[11]) if len(row) > 11 else "",
                "condition_class": safe_str(row[12]) if len(row) > 12 else "",
                "parameters": params,
                "tags": safe_str(row[15]) if len(row) > 15 else "",
                "ability_optimized": safe_str(row[16]) if len(row) > 16 else "",
                "suit": safe_str(row[17]) if len(row) > 17 else "",
            })
    return cards


def read_part2():
    """Read claud_raw_pt2_rough.txt (Part 2: TE, WR, DB, DL, LB) — 16 columns, no Suit."""
    cards = []
    with open(PART2_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if len(row) < 8 or not row[0].strip():
                continue
            cards.append({
                "name": safe_str(row[0]),
                "pos": safe_str(row[1]).upper(),
                "superstar": safe_str(row[2]),
                "passive": safe_str(row[3]),
                "run_mod": safe_int(row[4]),
                "short_mod": safe_int(row[5]),
                "deep_mod": safe_int(row[6]),
                "stamina": safe_int(row[7]),
                "grit": safe_int(row[8]) if len(row) > 8 else 0,
                "target": safe_str(row[9]) if len(row) > 9 else "",
                "bonus_amount": safe_str(row[10]) if len(row) > 10 else "",
                "affected_stat": safe_str(row[11]) if len(row) > 11 else "",
                "condition_class": safe_str(row[12]) if len(row) > 12 else "",
                "parameters": safe_str(row[13]) if len(row) > 13 else "",
                "oper": safe_str(row[14]) if len(row) > 14 else "",
                "tags": safe_str(row[15]) if len(row) > 15 else "",
                "ability_optimized": "",
                "suit": "",
            })
    return cards


# ── Write to Excel ───────────────────────────────────────────────────────────

def write_to_excel(all_cards):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Cards"]

    for i, card in enumerate(all_cards):
        row = i + 2  # row 1 is header
        card_id_num = i + 1
        card_id = format_card_id(card_id_num)
        pos = card["pos"]

        # Basic columns
        ws.cell(row, COL["card_id"], card_id)
        ws.cell(row, COL["Name"], card["name"])
        ws.cell(row, COL["pos"], pos)
        ws.cell(row, COL["Superstar"], card["superstar"])
        ws.cell(row, COL["passive"], card["passive"])
        ws.cell(row, COL["Suit"], card["suit"])
        ws.cell(row, COL["Stamina"], card["stamina"])
        ws.cell(row, COL["Grit"], card["grit"] if card["grit"] else "")

        # Stats — split by offense vs defense
        if pos in OFFENSIVE_POS:
            ws.cell(row, COL["RunBonus"], card["run_mod"])
            ws.cell(row, COL["ShortPass"], card["short_mod"])
            ws.cell(row, COL["DeepPass"], card["deep_mod"])
            ws.cell(row, COL["RunCov"], "")
            ws.cell(row, COL["ShortCov"], "")
            ws.cell(row, COL["DeepCov"], "")
        elif pos in DEFENSIVE_POS:
            ws.cell(row, COL["RunBonus"], "")
            ws.cell(row, COL["ShortPass"], "")
            ws.cell(row, COL["DeepPass"], "")
            ws.cell(row, COL["RunCov"], card["run_mod"])
            ws.cell(row, COL["ShortCov"], card["short_mod"])
            ws.cell(row, COL["DeepCov"], card["deep_mod"])
        else:
            ws.cell(row, COL["RunBonus"], card["run_mod"])
            ws.cell(row, COL["ShortPass"], card["short_mod"])
            ws.cell(row, COL["DeepPass"], card["deep_mod"])

        # Ability mapping
        a1, a2, manual_reason = map_ability(
            card_id, card["name"], pos, card["passive"],
            card["tags"], card.get("ability_optimized", "")
        )

        # Write a1 fields
        for key, val in a1.items():
            if key in COL:
                ws.cell(row, COL[key], val)

        # Write a2 fields
        for key, val in a2.items():
            if key in COL:
                ws.cell(row, COL[key], val)

        # Track for report
        if manual_reason:
            report_manual.append((card_id, card["name"], pos, manual_reason))

    wb.save(EXCEL_PATH)
    return len(all_cards)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Reading Part 1 (OL, QB, RB)...")
    part1 = read_part1()
    print(f"  Found {len(part1)} cards")

    print("Reading Part 2 (TE, WR, DB, DL, LB)...")
    part2 = read_part2()
    print(f"  Found {len(part2)} cards")

    all_cards = part1 + part2
    print(f"\nTotal cards: {len(all_cards)}")

    # Position breakdown
    from collections import Counter
    pos_counts = Counter(c["pos"] for c in all_cards)
    for pos in ["OL", "QB", "RB", "TE", "WR", "DL", "LB", "DB"]:
        print(f"  {pos}: {pos_counts.get(pos, 0)}")

    print(f"\nWriting to Excel: {EXCEL_PATH}")
    count = write_to_excel(all_cards)
    print(f"  Wrote {count} cards to rows 2-{count+1}")

    # ── Report ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 80)
    print("REPORT: Cards requiring manual review / new code")
    print("=" * 80)

    if report_manual:
        # Group by issue type
        charge_cards = [(c, n, p, r) for c, n, p, r in report_manual if "Charge" in r]
        discard_cards = [(c, n, p, r) for c, n, p, r in report_manual if "Discard" in r or "discard" in r.lower()]
        prevent_cards = [(c, n, p, r) for c, n, p, r in report_manual if "Prevent" in r or "prevent" in r.lower()]
        yardage_cards = [(c, n, p, r) for c, n, p, r in report_manual if "YardageGained" in r]
        seq_cards = [(c, n, p, r) for c, n, p, r in report_manual if "Sequential" in r or "Consecutive" in r or "stacking" in r.lower()]
        knockout_cards = [(c, n, p, r) for c, n, p, r in report_manual if "Knockout" in r or "knockout" in r.lower() or "stamina removal" in r.lower()]
        grit_cards = [(c, n, p, r) for c, n, p, r in report_manual if "Grit" in r or "grit" in r.lower()]
        slot_cards = [(c, n, p, r) for c, n, p, r in report_manual if "Slot" in r or "slot" in r.lower()]
        other_cards = [(c, n, p, r) for c, n, p, r in report_manual
                       if not any(kw in r for kw in ["Charge", "Discard", "discard", "Prevent", "prevent",
                                                      "YardageGained", "Sequential", "Consecutive",
                                                      "Knockout", "knockout", "stamina removal",
                                                      "Slot", "slot"])
                       and "Grit" not in r and "grit" not in r.lower()]

        def print_group(title, cards):
            if not cards:
                return
            print(f"\n--- {title} ({len(cards)} cards) ---")
            for cid, name, pos, reason in cards:
                print(f"  [{cid}] {name} ({pos}): {reason}")

        print_group("CHARGE MECHANICS (need EffectCharge wiring)", charge_cards)
        print_group("DISCARD-AS-COST (need EffectDiscardCard)", discard_cards)
        print_group("PREVENT LOSS (need EffectPreventLoss)", prevent_cards)
        print_group("YARDAGE GAINED CONDITION (need ConditionYardageGained)", yardage_cards)
        print_group("SEQUENTIAL / STACKING (need counter tracking)", seq_cards)
        print_group("KNOCKOUT / STAMINA REMOVAL", knockout_cards)
        print_group("GRIT MECHANICS", grit_cards)
        print_group("SLOT MANIPULATION", slot_cards)
        print_group("OTHER / UNCLASSIFIED", other_cards)

    auto_count = len(all_cards) - len(report_manual)
    print(f"\n--- SUMMARY ---")
    print(f"  Auto-mapped:  {auto_count} / {len(all_cards)} cards")
    print(f"  Need review:  {len(report_manual)} / {len(all_cards)} cards")
    print(f"  Excel saved:  {EXCEL_PATH}")


if __name__ == "__main__":
    main()
