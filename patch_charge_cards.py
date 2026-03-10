"""
Patches the simple charge cards into the Excel spreadsheet.
Run from the project root:
    python patch_charge_cards.py

Cards wired (simple: one stat bonus on threshold):
  00005 Maxwell Payne  (OL) - 20 helmets → +20 run
  00022 Tre Rostic     (OL) - 8 footballs → +5 run
  00100 Henry Walker   (DL) - 5 stars → +6 run coverage
  00113 William Ford   (DL) - 3 incomplete passes → +6 run coverage (next play)

Cards left MANUAL (bonus + secondary effect):
  00004 Andre Cross    (OL) - 10 footballs → +5 run AND restore 2 stamina
  00019 Tyson Rivers   (OL) - 5 stars → +4 run AND draw card
  00038 Kendrick Clovis(OL) - 10 helmets → next SP can't be sacked (EffectPreventLoss)
"""

import openpyxl
import os

XLSX_PATH = os.path.join("Assets", "Resources", "First and Long - card_exporter.xlsx")

COL_A1_TRIGGER    = 15
COL_A1_TARGET     = 16
COL_A1_FAIL_EVENT = 17
COL_A1_COND       = 18
COL_A1_COND_VAL   = 19
COL_A1_COND2      = 20
COL_A1_COND2_VAL  = 21
COL_A1_EFFECT     = 22
COL_A1_EFFECT_VAL = 23

COL_A2_TRIGGER    = 24
COL_A2_TARGET     = 25
COL_A2_FAIL_EVENT = 26
COL_A2_COND       = 27
COL_A2_COND_VAL   = 28
COL_A2_COND2      = 29
COL_A2_COND2_VAL  = 30
COL_A2_EFFECT     = 31
COL_A2_EFFECT_VAL = 32

def clear_a2(d):
    for c in [COL_A2_TRIGGER, COL_A2_TARGET, COL_A2_FAIL_EVENT, COL_A2_COND,
              COL_A2_COND_VAL, COL_A2_COND2, COL_A2_COND2_VAL, COL_A2_EFFECT, COL_A2_EFFECT_VAL]:
        d[c] = None

PATCHES = {}

# Maxwell Payne (OL): 20 helmets → +20 run on current run
p = {}
p[COL_A1_TRIGGER]    = "OnRun"
p[COL_A1_TARGET]     = "Self"
p[COL_A1_FAIL_EVENT] = None
p[COL_A1_COND]       = None
p[COL_A1_COND_VAL]   = None
p[COL_A1_COND2]      = None
p[COL_A1_COND2_VAL]  = None
p[COL_A1_EFFECT]     = "Charge"
p[COL_A1_EFFECT_VAL] = "Helmet|20|RunBonus|20|1"
clear_a2(p)
PATCHES["00005"] = p   # verify card_id matches Maxwell Payne

# Tre Rostic (OL): 8 footballs → +5 run
p = {}
p[COL_A1_TRIGGER]    = "OnRun"
p[COL_A1_TARGET]     = "Self"
p[COL_A1_FAIL_EVENT] = None
p[COL_A1_COND]       = None
p[COL_A1_COND_VAL]   = None
p[COL_A1_COND2]      = None
p[COL_A1_COND2_VAL]  = None
p[COL_A1_EFFECT]     = "Charge"
p[COL_A1_EFFECT_VAL] = "Football|8|RunBonus|5|1"
clear_a2(p)
PATCHES["00032"] = p   # verify card_id matches Tre Rostic

# Henry Walker (DL): 5 stars → +6 run coverage on current run
p = {}
p[COL_A1_TRIGGER]    = "OnRun"
p[COL_A1_TARGET]     = "Self"
p[COL_A1_FAIL_EVENT] = None
p[COL_A1_COND]       = None
p[COL_A1_COND_VAL]   = None
p[COL_A1_COND2]      = None
p[COL_A1_COND2_VAL]  = None
p[COL_A1_EFFECT]     = "Charge"
p[COL_A1_EFFECT_VAL] = "Star|5|RunCoverage|6|1"
clear_a2(p)
PATCHES["00100"] = p   # verify card_id matches Henry Walker

# William Ford (DL): 3 incomplete passes → +6 run coverage, duration=2 (carries into next play)
p = {}
p[COL_A1_TRIGGER]    = "OnPass"
p[COL_A1_TARGET]     = "Self"
p[COL_A1_FAIL_EVENT] = None
p[COL_A1_COND]       = "LastPassIncomplete"
p[COL_A1_COND_VAL]   = None
p[COL_A1_COND2]      = None
p[COL_A1_COND2_VAL]  = None
p[COL_A1_EFFECT]     = "Charge"
p[COL_A1_EFFECT_VAL] = "None|3|RunCoverage|6|2"
clear_a2(p)
PATCHES["00113"] = p   # verify card_id matches William Ford

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: File not found: {XLSX_PATH}")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Print header row to verify column indices
    headers = [ws.cell(row=1, column=c).value for c in range(1, 33)]
    print("Col 15-23:", headers[14:23])
    print()

    # Build card_id -> row mapping; also print names to verify IDs
    id_to_row = {}
    id_to_name = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=False):
        id_cell, name_cell = row[0], row[1]
        if id_cell.value is not None:
            cid = str(id_cell.value).strip()
            id_to_row[cid] = id_cell.row
            id_to_name[cid] = name_cell.value

    # Report which IDs were found
    for cid in PATCHES:
        name = id_to_name.get(cid, "NOT FOUND")
        print(f"  {cid} -> {name}")
    print()

    updated = []
    missing = []
    for card_id, cols in PATCHES.items():
        if card_id not in id_to_row:
            missing.append(card_id)
            continue
        row_idx = id_to_row[card_id]
        for col_idx, value in cols.items():
            ws.cell(row=row_idx, column=col_idx).value = value
        updated.append(f"{card_id} ({id_to_name.get(card_id, '?')})")

    wb.save(XLSX_PATH)
    print(f"Saved {XLSX_PATH}")
    print(f"Updated ({len(updated)}): {', '.join(updated)}")
    if missing:
        print(f"NOT FOUND ({len(missing)}): {', '.join(missing)}")
        print("Check card IDs in Excel column A vs the IDs in this script.")

if __name__ == "__main__":
    main()
